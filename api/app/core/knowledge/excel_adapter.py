"""Structure-preserving Excel workbook adapter for the canonical Document IR."""

from __future__ import annotations

import io
from collections.abc import Iterable
from typing import Any

from app.core.knowledge.ir import BlockKind, DocumentBlock, DocumentIR, SourceAnchor


def _display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def _row_text(headers: list[str], values: list[Any], row_number: int) -> str:
    rendered = [_display(value) for value in values]
    pairs = []
    for index, value in enumerate(rendered):
        if not value:
            continue
        header = headers[index] if index < len(headers) and headers[index] else f"列{index + 1}"
        pairs.append(f"{header}: {value}")
    return f"第{row_number}行 | " + " | ".join(pairs)


def _sheet_blocks(
    rows: Iterable[list[Any]],
    *,
    sheet_name: str,
    sheet_index: int,
    document_id: str,
    version_id: str,
    start_order: int,
) -> list[DocumentBlock]:
    materialized = [list(row) for row in rows]
    nonempty = [row for row in materialized if any(_display(value) for value in row)]
    anchor = SourceAnchor(document_id=document_id, version_id=version_id, parser="excel", parser_version="1")
    blocks = [
        DocumentBlock(
            block_id=f"{document_id}:sheet:{sheet_index}:title",
            kind=BlockKind.HEADING,
            content=f"工作表：{sheet_name}",
            anchor=anchor,
            order=start_order,
            section_path=(sheet_name,),
            metadata={"sheet_name": sheet_name, "sheet_index": sheet_index},
        )
    ]
    if not nonempty:
        return blocks
    width = max(len(row) for row in nonempty)
    first = nonempty[0] + [None] * (width - len(nonempty[0]))
    headers = [_display(value) or f"列{index + 1}" for index, value in enumerate(first)]
    table_id = f"{document_id}:sheet:{sheet_index}:table"
    blocks.append(
        DocumentBlock(
            block_id=table_id,
            kind=BlockKind.TABLE,
            content="表头 | " + " | ".join(headers),
            anchor=anchor,
            order=start_order + 1,
            section_path=(sheet_name,),
            logical_table_id=table_id,
            metadata={"sheet_name": sheet_name, "row_start": 1, "row_end": len(nonempty)},
        )
    )
    for offset, row in enumerate(nonempty[1:], start=2):
        padded = row + [None] * (width - len(row))
        text = _row_text(headers, padded, offset)
        if text.endswith("| "):
            continue
        blocks.append(
            DocumentBlock(
                block_id=f"{table_id}:row:{offset}",
                kind=BlockKind.TABLE_ROW,
                content=text,
                anchor=anchor,
                order=start_order + offset,
                section_path=(sheet_name,),
                parent_block_id=table_id,
                logical_table_id=table_id,
                metadata={"sheet_name": sheet_name, "row_number": offset, "headers": headers},
            )
        )
    return blocks


def excel_to_ir(
    content: bytes,
    *,
    file_ext: str,
    document_id: str,
    version_id: str,
    title: str,
) -> DocumentIR:
    ext = file_ext.lower()
    blocks: list[DocumentBlock] = []
    sheet_names: list[str] = []
    if ext in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        try:
            for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
                sheet_names.append(sheet.title)
                rows = ([cell.value for cell in row] for row in sheet.iter_rows())
                blocks.extend(_sheet_blocks(rows, sheet_name=sheet.title, sheet_index=sheet_index,
                    document_id=document_id, version_id=version_id, start_order=len(blocks) * 1000))
        finally:
            workbook.close()
    elif ext == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        try:
            for sheet_index, sheet_name in enumerate(workbook.sheet_names(), start=1):
                sheet_names.append(sheet_name)
                sheet = workbook.sheet_by_name(sheet_name)
                rows = (sheet.row_values(index) for index in range(sheet.nrows))
                blocks.extend(_sheet_blocks(rows, sheet_name=sheet_name, sheet_index=sheet_index,
                    document_id=document_id, version_id=version_id, start_order=len(blocks) * 1000))
        finally:
            workbook.release_resources()
    else:
        raise ValueError(f"unsupported Excel extension: {file_ext}")
    return DocumentIR(
        document_id=document_id,
        version_id=version_id,
        title=title,
        blocks=blocks,
        metadata={"format": "excel", "sheet_names": sheet_names, "sheet_count": len(sheet_names)},
    )
